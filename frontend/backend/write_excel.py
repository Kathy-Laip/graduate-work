import pandas as pd
import numpy as np
from openpyxl import load_workbook
import mysql.connector


DEFAULT_FILENAME = 'Направления.xlsx'
DEFAULT_SHEET = 'Лист1'

# # подключение к базе данных
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password='regopi09'
)
mycursor = mydb.cursor()

def get_columns_auto(name_db, name_table):
    name_db += '.' + name_table

    # запрашиваем у базы данных для таблицы название колонок, чтобы не прописывать в ручную
    mycursor.execute('show columns from {};'.format(name_db))
    name_columns = mycursor.fetchall()
    columns = [name_columns[i][0] for i in range(1,len(name_columns))] # кортеж о данных таблицы
    columns = ', '.join(columns)
    return columns

def get_columns(name_db, name_table):
    name_db += '.' + name_table

    # запрашиваем у базы данных для таблицы название колонок, чтобы не прописывать в ручную
    mycursor.execute('show columns from {};'.format(name_db))
    name_columns = mycursor.fetchall()
    columns = [name_columns[i][0] for i in range(0,len(name_columns))] # кортеж о данных таблицы
    columns = ', '.join(columns)
    return columns

def select(name_db, name_table):
    name_db += '.' + name_table
    mycursor.execute('select * from {};'.format(name_db))
    data = mycursor.fetchall()
    return data

def insert_into(name_db, data, columns, count):
    print(name_db, columns, count)
    for i in range(len(data)):
        count = ''.join(count)
        mycursor.execute('INSERT INTO {}({}) VALUES({});'.format(name_db, columns, count), data[i])
        mydb.commit()
    return 'ok'

def delete_data(name_db, name_table):
    name_db += '.' + name_table
    mycursor.execute('DELETE FROM {};'.format(name_db))
    mydb.commit()
    return 'all data has been deleted'

# функция для считывания данных из excel-файла в оперативную память программы
# filename - путь до файла, который хотим открыть
# sheet_name - название excel листа, который хотим считывать
def read_excel(filename, sheet_name=DEFAULT_SHEET):
    wb = load_workbook(filename=filename)
    timetable = wb[sheet_name] # получаем лист с названием из переменной sheet_name
    return (wb, timetable)

def save_changes(workbook, filename):
    workbook.save(filename)

def get_cell(sheet, column, row):
    return sheet.cell(row=row, column=column).value

if __name__ == '__main__':
    wb, strs = read_excel(DEFAULT_FILENAME)

    strs_arr = [] # все строки из таблицы excel

    for row in strs.iter_rows(values_only=True):
        str_arr = []
        for cell in row:
            str_arr.append(cell)
        strs_arr.append(str_arr) # каждую строку добавляем в общаг
    for i in range(len(strs_arr)):
        print(strs_arr[i]) # выводим строки

    data_course = select('schedules', 'courses')  # получаем все курсы и их индексы
    data_course_dict = {} # словарь для того чтобы хранить индекс строки принадлежности курса, необходимы для того чтобы вставлять данные о направлениях
    print(data_course)
    for i in range(len(data_course)): # составляем словарь
        data_course_dict[data_course[i][1]] = data_course[i][0]
    
    direction_columns = get_columns_auto('schedules', 'direction') # получаем колонки таблицы с направлениями, для вставки
    delete_data('schedules', 'direction')

    print(data_course_dict)
    data_direction = set()
    for i in range(1, len(strs_arr)):
        data_direction.add((data_course_dict[strs_arr[i][0]], strs_arr[i][1])) # обработка данных для вставки
    print(data_direction)
    data_direction = list(data_direction)

    count = ['%s' for i in range(len(data_direction[0]))] # форматируем колонки, чтобы подставить дальше данные
    count = ', '.join(count)
    
    insert_into('schedules.direction', data_direction, direction_columns, count) # вставка данных 

    data_dir = select('schedules', 'direction')  # получаем все направления и их индексы
    data_dir_dict = {} #направления
    for i in range(len(data_dir)): # составляем словарь
        data_dir_dict[data_dir[i][2]] = data_dir[i][0]
    print(data_dir)

    classes_columns = get_columns_auto('schedules', 'classes') # получаем колонки таблицы с направлениями, для вставки
    print(classes_columns)

    data_classes = []
    for i in range(1, len(strs_arr)):
        data_classes.append((data_dir_dict[strs_arr[i][1]], strs_arr[i][2], strs_arr[i][3]))
    print(data_classes)

    count_cl = ['%s, ' for i in range(len(data_classes[0])-1)] # форматируем колонки, чтобы подставить дальше данные
    count_cl.append('%s')
    

    insert_into('schedules.classes', data_classes, classes_columns, count_cl)
